
i = 9 # delete tha
import openpyxl
# WEEK NB IS SET TO += 15 !!
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import date

alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

wb = openpyxl.load_workbook('weekly_review.xlsx')


# grab the active worksheet
ws = wb.active
# creating the week in the lines

i_for_content = 32
i_for_cell = 2
while i_for_content < 53:
    i_for_cell = str(i_for_cell)
    cell = "A" + i_for_cell
    i_for_cell = int(i_for_cell)
    ws[cell] = i_for_content
    i_for_cell += 1
    i_for_content += 1




# put the habits on the 1st line
habits = ["Carnegie", "Drink", "Use transportation", "Intresting TgB", "Beautiful notebooks", "fr./g.", "c*n*", "passionate + part.", "Music and places", "grat", "home: f., (G), in advance, autotests, efficace..."]
index_for_habits = 0
index_for_alphabet = 0
while index_for_habits < len(habits):
    cell = str(alphabet[index_for_alphabet + 1].upper()) + "1"
    ws[cell] = habits[index_for_habits]
    index_for_alphabet += 1
    index_for_habits += 1

ws["B1"] = habits[0]
ws["C1"] = habits[1]
ws["D1"] = habits[2]
ws["E1"] = habits[3]
ws["F1"] = habits[4]
ws["G1"] = habits[5]
ws["H1"] = habits[6]
ws["I1"] = habits[7]
ws["J1"] = habits[8]
ws["K1"] = habits[9]

ws["N1"] = "home: f.,  (G), in advance, autotests, wllm..."




week_number = weekNumber = date.today().isocalendar()[1] 
carnegie = input(f"{habits[0]}  : ")
drink = input(f"{habits[1]}  : ")
use_transportation = input(f"{habits[2]}  : ")
tgb = input(f"{habits[3]}  : ")
notebooks = input(f"{habits[4]}  : ")
fr_g = input(f"{habits[5]}  : ")
c_n = input(f"{habits[6]}  : ")
passionate = input(f"{habits[7]}  : ")
music = input(f"{habits[8]}  : ")
grat = input(f"{habits[9]}  : ")
home = input(f"{habits[10]}  : ")# habits = ["Carnegie", "Drink", "Use transportation", "S: podcasts, videos, P...", "Intresting TgB", "Beautiful notebooks", "fr./g.", "c*n*", "passionate + part.", "Music and places", "grat", "home: f., (G), in advance, autotests, efficace..."]

todays_line = int(week_number) #- 33 # to change as well
evaluations = [carnegie, drink, use_transportation,  tgb, notebooks,  fr_g, c_n, passionate, music, grat, home]

index = 0
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
orangeFill = PatternFill(start_color='F79D00',
                   end_color='F79D00',
                   fill_type='solid')
blueFill = PatternFill(start_color='0A9EFF',
                   end_color='0A9EFF',
                   fill_type='solid')
greenFill = PatternFill(start_color='98F50E', 
                   end_color='98F50E',
                   fill_type='solid')

while index < len(evaluations):
    evaluation = evaluations[index]
    cell_to_handle = alphabet[index + 1].upper() + str(todays_line - 30 )
    print(cell_to_handle)
    if evaluation == "1":
       ws[cell_to_handle].fill = redFill
    elif evaluation == "2":
        ws[cell_to_handle].fill = orangeFill
    elif evaluation == "3":
        ws[cell_to_handle].fill = blueFill
    elif evaluation == "4":
        ws[cell_to_handle].fill = greenFill
        print("cell to handle", cell_to_handle)
    index += 1
wb.save("weekly_review.xlsx")
# create new ws for txts habits

